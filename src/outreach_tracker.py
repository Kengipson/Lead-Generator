"""
Append normalized leads into the "Outreach Tracker" sheet of an existing
outreach-tracker workbook (e.g. kenai-outreach-tracker.xlsx), skipping
businesses that are already in the sheet.

This does not create or redesign the workbook -- it expects one that already
has an "Outreach Tracker" sheet with this header row:

    Business Name | Industry | Contact Name | Phone / Email |
    Source (Chamber, Maps, Referral, etc.) | Date Contacted |
    Free Sample Sent? (Y/N) | Sample Sent Date | Response (Y/N/Pending) |
    Follow-Up Date | Status | Notes
"""

from __future__ import annotations

import re
from copy import copy
from pathlib import Path
from typing import Any

import openpyxl

SHEET_NAME = "Outreach Tracker"

# Column order in the tracker sheet (1-indexed to match openpyxl).
COLUMNS = [
    "Business Name",
    "Industry",
    "Contact Name",
    "Phone / Email",
    "Source (Chamber, Maps, Referral, etc.)",
    "Date Contacted",
    "Free Sample Sent? (Y/N)",
    "Sample Sent Date",
    "Response (Y/N/Pending)",
    "Follow-Up Date",
    "Status",
    "Notes",
]
COL_NAME = 1
COL_PHONE = 4


def _normalize_name(name: str) -> str:
    return " ".join((name or "").strip().lower().split())


def _normalize_phone(phone: str) -> str:
    """Keep digits only, so formatting differences don't defeat dedup."""
    return re.sub(r"\D", "", phone or "")


def _existing_keys(ws) -> tuple[set[str], set[str]]:
    """Collect normalized business names and phone numbers already in the sheet."""
    names: set[str] = set()
    phones: set[str] = set()
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        name_val = row[COL_NAME - 1].value
        phone_val = row[COL_PHONE - 1].value
        if name_val:
            names.add(_normalize_name(str(name_val)))
        if phone_val:
            norm_phone = _normalize_phone(str(phone_val))
            if norm_phone:
                phones.add(norm_phone)
    return names, phones


def _next_blank_row(ws) -> int:
    """First row (from row 2 on) whose Business Name cell is empty."""
    row = 2
    while ws.cell(row=row, column=COL_NAME).value not in (None, ""):
        row += 1
    return row


def _apply_template_style(ws, row: int, template_row: int) -> None:
    """Copy font/fill/border/alignment/number_format from template_row onto row."""
    for col in range(1, len(COLUMNS) + 1):
        src = ws.cell(row=template_row, column=col)
        dst = ws.cell(row=row, column=col)
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format


def _build_notes(industry: str, location: str, high_priority_no_website: bool) -> str:
    notes = f"Found via Maps search - {industry} in {location}"
    if high_priority_no_website:
        notes += " - no website (high priority)"
    return notes


def append_leads_to_tracker(
    leads: list[dict[str, Any]],
    industry: str,
    location: str,
    tracker_path: str | Path,
) -> dict[str, int]:
    """Append new leads to the Outreach Tracker sheet, skipping duplicates.

    A lead is considered a duplicate if its business name OR phone number
    (digits only) already appears anywhere in the sheet.

    Returns a summary dict: {"appended": N, "skipped_duplicates": N}.
    """
    tracker_path = Path(tracker_path)
    if not tracker_path.exists():
        raise FileNotFoundError(f"Outreach tracker file not found: {tracker_path}")

    wb = openpyxl.load_workbook(tracker_path)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(
            f'"{SHEET_NAME}" sheet not found in {tracker_path}. '
            f"Found sheets: {wb.sheetnames}"
        )
    ws = wb[SHEET_NAME]

    existing_names, existing_phones = _existing_keys(ws)

    appended = 0
    skipped = 0
    template_row = max(3, ws.max_row) if ws.max_row >= 3 else 2

    for lead in leads:
        name = (lead.get("name") or "").strip()
        phone = (lead.get("phone") or "").strip()
        norm_name = _normalize_name(name)
        norm_phone = _normalize_phone(phone)

        is_duplicate = (norm_name and norm_name in existing_names) or (
            norm_phone and norm_phone in existing_phones
        )
        if is_duplicate:
            skipped += 1
            continue

        row = _next_blank_row(ws)
        if row > ws.max_row:
            # Ran out of pre-formatted template rows -- extend, matching style.
            _apply_template_style(ws, row, template_row)

        values = {
            "Business Name": name,
            "Industry": industry,
            "Contact Name": None,
            "Phone / Email": phone,
            "Source (Chamber, Maps, Referral, etc.)": "Google Maps",
            "Date Contacted": None,
            "Free Sample Sent? (Y/N)": None,
            "Sample Sent Date": None,
            "Response (Y/N/Pending)": None,
            "Follow-Up Date": None,
            "Status": "New Lead",
            "Notes": _build_notes(industry, location, lead.get("high_priority_no_website", False)),
        }
        for col_index, col_name in enumerate(COLUMNS, start=1):
            ws.cell(row=row, column=col_index).value = values[col_name]

        if norm_name:
            existing_names.add(norm_name)
        if norm_phone:
            existing_phones.add(norm_phone)
        appended += 1

    wb.save(tracker_path)
    return {"appended": appended, "skipped_duplicates": skipped}
